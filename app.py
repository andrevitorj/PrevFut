import streamlit as st
import requests
import json
import pandas as pd
import numpy as np
from scipy.stats import poisson, norm
import openpyxl
from datetime import datetime
import os
from dotenv import load_dotenv
import logging
import matplotlib.pyplot as plt
import seaborn as sns

# Configurar logging para depuração (não será exibido na interface)
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Carregar variáveis de ambiente
load_dotenv()
API_KEY = os.getenv("API_FOOTBALL_KEY")

# Configuração da API
API_BASE_URL = "https://v3.football.api-sports.io"
HEADERS = {
    "x-apisports-key": API_KEY
}

# Verificar chave da API
if not API_KEY:
    st.error("Chave da API não encontrada. Configure 'API_FOOTBALL_KEY' nos secrets do Streamlit.")
    st.stop()

# Função para testar a chave da API
def test_api_key():
    url = f"{API_BASE_URL}/status"
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code == 200:
            return True
        else:
            st.error(f"Erro ao testar chave: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        st.error(f"Erro ao testar chave: {str(e)}")
        return False

# Função para reformatar a coluna "Club / Country"
def reformat_club_country(club_country):
    countries = [
        "Italy", "France", "England", "Spain", "Germany", "Portugal", 
        "Netherlands", "Turkey", "Brazil", "Argentina", "Belgium", "Scotland"
    ]
    for country in countries:
        if club_country.endswith(country):
            team_name = club_country[:len(club_country) - len(country)].strip()
            return f"{team_name} ({country})"
    return club_country

# Função para buscar os pontos Elo e calcular o peso
def get_team_elo_and_weight(team_name, ratings_df):
    if ratings_df is None:
        return None, 0.8  # Pontos Elo não encontrados, peso padrão
    # Remove o país do nome do time para facilitar a busca
    team_row = ratings_df[ratings_df["Club / Country"].str.contains(team_name, case=False, na=False)]
    if not team_row.empty:
        points = float(team_row["Points"].iloc[0])
        weight = (points / 2000) ** 2  # Fórmula: (points/2000)^2
        return points, weight
    return None, 0.8  # Pontos Elo não encontrados, peso padrão

# Função para calcular o peso a partir dos pontos Elo
def calculate_weight_from_elo(points):
    return (points / 2000) ** 2  # Fórmula: (points/2000)^2

# Função para buscar times
def search_team(team_name):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return []
    url = f"{API_BASE_URL}/teams"
    params = {"search": team_name}
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        if response.status_code == 200:
            return response.json().get("response", [])
        else:
            st.error(f"Erro na API: {response.status_code} - {response.text}")
            return []
    except Exception as e:
        st.error(f"Erro ao buscar times: {str(e)}")
        return []

# Função para buscar jogos passados
def get_team_games(team_id, season, home=True, limit=20, neutral=False):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return []
    url = f"{API_BASE_URL}/fixtures"
    params = {
        "team": team_id,
        "season": season,
        "last": limit,
        "status": "FT"  # Apenas jogos finalizados
    }
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        if response.status_code == 200:
            data = response.json()
            games = data.get("response", [])
            if neutral:
                return games  # retorna todos os jogos
            filtered_games = [
                game for game in games
                if (home and game["teams"]["home"]["id"] == team_id) or
                   (not home and game["teams"]["away"]["id"] == team_id)
            ]
            if not filtered_games and season == 2025:
                st.warning(f"Nenhum jogo encontrado para temporada {season}. Tentando temporada {season-1}...")
                params["season"] = season - 1
                response = requests.get(url, headers=HEADERS, params=params)
                if response.status_code == 200:
                    data = response.json()
                    games = data.get("response", [])
                    if neutral:
                        return games
                    filtered_games = [
                        game for game in games
                        if (home and game["teams"]["home"]["id"] == team_id) or
                           (not home and game["teams"]["away"]["id"] == team_id)
                    ]
            return filtered_games
        st.error(f"Erro ao buscar jogos: {response.status_code} - {response.text}")
        return []
    except Exception as e:
        st.error(f"Erro ao buscar jogos: {str(e)}")
        return []


# Função para buscar o próximo jogo entre dois times
def find_next_fixture(team_a_id, team_b_id, season):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return None
    url = f"{API_BASE_URL}/fixtures"
    params = {
        "team": team_a_id,
        "season": season,
        "next": 20
    }
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        logger.debug(f"Resposta da API para buscar fixture: {response.status_code} - {response.text}")
        if response.status_code == 200:
            data = response.json()
            games = data.get("response", [])
            for game in games:
                if (game["teams"]["home"]["id"] == team_a_id and game["teams"]["away"]["id"] == team_b_id) or \
                   (game["teams"]["home"]["id"] == team_b_id and game["teams"]["away"]["id"] == team_a_id):
                    fixture_id = game["fixture"]["id"]
                    logger.debug(f"Fixture encontrado: {fixture_id} para {team_a_id} vs {team_b_id}")
                    return fixture_id
            logger.debug(f"Nenhum jogo futuro encontrado entre {team_a_id} e {team_b_id} na temporada {season}")
            st.warning("Nenhum jogo futuro encontrado entre os times selecionados na temporada especificada.")
            return None
        st.error(f"Erro ao buscar próximo jogo: {response.status_code} - {response.text}")
        return None
    except Exception as e:
        st.error(f"Erro ao buscar próximo jogo: {str(e)}")
        return None

# Função para buscar estatísticas de um jogo
def get_game_stats(fixture_id):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return []
    url = f"{API_BASE_URL}/fixtures/statistics"
    params = {"fixture": fixture_id}
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        if response.status_code == 200:
            return response.json().get("response", [])
        st.error(f"Erro ao buscar estatísticas: {response.status_code} - {response.text}")
        return []
    except Exception as e:
        st.error(f"Erro ao buscar estatísticas: {str(e)}")
        return []

# Função para buscar competições de um time
def get_team_leagues(team_id, season):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return []
    url = f"{API_BASE_URL}/leagues"
    params = {"team": team_id, "season": season}
    try:
        response = requests.get(url, headers=HEADERS, params=params)
        if response.status_code == 200:
            return response.json().get("response", [])
        st.error(f"Erro ao buscar ligas: {response.status_code} - {response.text}")
        return []
    except Exception as e:
        st.error(f"Erro ao buscar ligas: {str(e)}")
        return []

# Função para calcular médias (feitas e sofridas) com número de partidas
def calculate_averages(games, team_id, season, team_weight, opponent_weights=None):
    if opponent_weights is None:
        opponent_weights = {}

    stats = {
        "goals_scored": [], "goals_conceded": [],
        "shots": [], "shots_conceded": [],
        "shots_on_target": [], "shots_on_target_conceded": [],
        "corners": [], "corners_conceded": [],
        "possession": [], "possession_conceded": [],
        "offsides": [], "offsides_conceded": [],
        "fouls_committed": [], "fouls_suffered": [],
        "yellow_cards": [], "yellow_cards_conceded": [],
        "red_cards": [], "red_cards_conceded": [],
        "passes_accurate": [], "passes_accurate_conceded": [],
        "passes_missed": [], "passes_missed_conceded": [],
        "xg": [], "xga": [],
        "free_kicks": [], "free_kicks_conceded": []
    }
    adjusted_values = {k: [] for k in stats.keys()}
    game_counts = {k: 0 for k in stats.keys()}

    stat_mapping = {
        "total shots": "shots",
        "shots on goal": "shots_on_target",
        "corner kicks": "corners",
        "ball possession": "possession",
        "offsides": "offsides",
        "fouls": "fouls_committed",
        "yellow cards": "yellow_cards",
        "red cards": "red_cards",
        "passes accurate": "passes_accurate",
        "passes": "passes_missed",
        "expected goals": "xg",
        "free kicks": "free_kicks"
    }
    stat_mapping_opponent = {
        "total shots": "shots_conceded",
        "shots on goal": "shots_on_target_conceded",
        "corner kicks": "corners_conceded",
        "ball possession": "possession_conceded",
        "offsides": "offsides_conceded",
        "fouls": "fouls_suffered",
        "yellow cards": "yellow_cards_conceded",
        "red cards": "red_cards_conceded",
        "passes accurate": "passes_accurate_conceded",
        "passes": "passes_missed_conceded",
        "expected goals": "xga",
        "expected goals against": "xga",
        "free kicks": "free_kicks_conceded"
    }

    for game in games:
        game_stats = get_game_stats(game["fixture"]["id"])
        team_data = {k: 0 for k in stats.keys()}
        has_stats_for_team = False
        has_stats_for_opponent = False
        is_home = game["teams"]["home"]["id"] == team_id
        team_data["goals_scored"] = game["goals"]["home" if is_home else "away"] or 0
        team_data["goals_conceded"] = game["goals"]["away" if is_home else "home"] or 0

        # Determinar o nome do adversário
        opponent_name = game["teams"]["away"]["name"] if is_home else game["teams"]["home"]["name"]
        # Obter o peso do adversário do session_state ou usar o padrão
        fixture_id = str(game["fixture"]["id"])
        opponent_weight = opponent_weights.get(fixture_id, 0.8)  # Peso padrão 0.8 se não ajustado

        stats_available = {k: False for k in stats.keys()}

        if game_stats:
            team_stats = next((s for s in game_stats if s["team"]["id"] == team_id), None)
            opponent_stats = next((s for s in game_stats if s["team"]["id"] != team_id), None)
            if team_stats:
                has_stats_for_team = True
                for stat in team_stats["statistics"]:
                    stat_type = stat["type"].lower()
                    value = stat["value"]
                    if value is None:
                        value = 0
                    elif stat_type == "ball possession" and isinstance(value, str):
                        value = value.replace("%", "").strip()
                        value = float(value) if value else 0.0
                    else:
                        try:
                            value = float(value) if value else 0.0
                        except (ValueError, TypeError):
                            value = 0.0
                    if stat_type in stat_mapping:
                        mapped_stat = stat_mapping[stat_type]
                        team_data[mapped_stat] = value
                        stats_available[mapped_stat] = True
                    if stat_type == "passes":
                        team_data["passes_missed"] = (value or 0) - (team_data["passes_accurate"] or 0)
                        stats_available["passes_missed"] = True

            if opponent_stats:
                has_stats_for_opponent = True
                for stat in opponent_stats["statistics"]:
                    stat_type = stat["type"].lower()
                    value = stat["value"]
                    if value is None:
                        value = 0
                    elif stat_type == "ball possession" and isinstance(value, str):
                        value = value.replace("%", "").strip()
                        value = float(value) if value else 0.0
                    else:
                        try:
                            value = float(value) if value else 0.0
                        except (ValueError, TypeError):
                            value = 0.0
                    if stat_type in stat_mapping_opponent:
                        mapped_stat = stat_mapping_opponent[stat_type]
                        team_data[mapped_stat] = value
                        stats_available[mapped_stat] = True
                    if stat_type == "passes":
                        team_data["passes_missed_conceded"] = (value or 0) - (team_data["passes_accurate_conceded"] or 0)
                        stats_available["passes_missed_conceded"] = True

        for key in stats:
            stats[key].append(team_data[key])
            if key in ["goals_scored", "goals_conceded"]:
                game_counts[key] += 1
            elif stats_available[key] and (has_stats_for_team or has_stats_for_opponent):
                game_counts[key] += 1

        for key in adjusted_values:
            if "conceded" in key or "suffered" in key:
                adjusted_values[key].append(team_data[key] / max(opponent_weight, 0.1))
            else:
                adjusted_values[key].append(team_data[key] * opponent_weight)

    simple_averages = {k: np.mean(v[:game_counts[k]]) if game_counts[k] > 0 else 0 for k, v in stats.items()}
    adjusted_averages = {k: np.mean(v[:game_counts[k]]) if game_counts[k] > 0 else 0 for k, v in adjusted_values.items()}
    return simple_averages, adjusted_averages, game_counts, team_weight, stats, adjusted_values

# Função para prever estatísticas (modificada para calcular apenas estatísticas favoráveis)
def predict_stats(team_a_simple, team_a_adjusted, team_b_simple, team_b_adjusted, team_a_counts, team_b_counts, team_a_weight, team_b_weight):
    if team_a_adjusted["goals_scored"] == 0 or team_b_adjusted["goals_scored"] == 0:
        st.warning("Não há dados suficientes de gols para previsões estatísticas confiáveis.")
        return {}, {}, {}

    predicted_stats = {}
    confidence_intervals = {}
    predicted_counts = {}

    # Apenas estatísticas favoráveis
    favorable_stats = ["goals_scored", "shots", "shots_on_target", "corners", "possession", "offsides", "fouls_committed", "yellow_cards", "red_cards", "passes_accurate", "passes_missed", "xg", "free_kicks"]
    
    stat_pairs = {
        "goals_scored": "goals_conceded",
        "shots": "shots_conceded",
        "shots_on_target": "shots_on_target_conceded",
        "corners": "corners_conceded",
        "possession": "possession_conceded",
        "offsides": "offsides_conceded",
        "fouls_committed": "fouls_suffered",
        "yellow_cards": "yellow_cards_conceded",
        "red_cards": "red_cards_conceded",
        "passes_accurate": "passes_accurate_conceded",
        "passes_missed": "passes_missed_conceded",
        "xg": "xga",
        "free_kicks": "free_kicks_conceded"
    }

    for stat in favorable_stats:
        opposite_stat = stat_pairs[stat]
        
                # Previsão para Time A com a nova fórmula
        if team_a_weight != 0:  # Evitar divisão por zero
            a_pred = (team_a_adjusted[stat] + team_b_adjusted[opposite_stat]) / (2 * (team_b_weight / team_a_weight))
        else:
            a_pred = (team_a_adjusted[stat] + team_b_adjusted[opposite_stat]) / 2  # Caso team_a_weight seja 0, evita erro

        # Previsão para Time B com a nova fórmula
        if team_b_weight != 0:  # Evitar divisão por zero
            b_pred = (team_b_adjusted[stat] + team_a_adjusted[opposite_stat]) / (2 * (team_a_weight / team_b_weight))
        else:
            b_pred = (team_b_adjusted[stat] + team_a_adjusted[opposite_stat]) / 2  # Caso team_b_weight seja 0, evita erro
        
        predicted_stats[stat] = {"team_a": a_pred, "team_b": b_pred}

        # Aproximação para intervalo de confiança usando Poisson para gols
        # Número de partidas combinado
        count = (team_a_counts[stat] + team_b_counts[opposite_stat]) // 2
        
        # Desvio padrão aproximado (usando Poisson para gols, ou média para outras stats)
        if stat == "goals_scored":
            std_a = np.sqrt(a_pred)  # Aproximação Poisson
            std_b = np.sqrt(b_pred)
        else:
            # Para outras estatísticas, usamos uma aproximação simples (não temos dados brutos)
            std_a = a_pred / 2  # Aproximação arbitrária
            std_b = b_pred / 2
        
        z = norm.ppf(0.925)  # Para 85% de confiança
        error_a = std_a / np.sqrt(count) if count > 0 else 0
        error_b = std_b / np.sqrt(count) if count > 0 else 0
        
        confidence_intervals[stat] = {
            "team_a": (max(0, a_pred - z * error_a), a_pred + z * error_a),
            "team_b": (max(0, b_pred - z * error_b), b_pred + z * error_b)
        }
        predicted_counts[stat] = count

    return predicted_stats, confidence_intervals, predicted_counts

# Função para prever placar
def predict_score(team_a_adjusted, team_b_adjusted, team_a_weight, team_b_weight, team_a_name, team_b_name):
    if team_a_adjusted["goals_scored"] == 0 or team_b_adjusted["goals_scored"] == 0:
        st.warning("Não há dados suficientes de gols para prever o placar.")
        return {"score": "N/A", "probs": {"win": 0, "draw": 0, "loss": 0}, "ci": {"team_a": (0, 0), "team_b": (0, 0)}, "total_goals_prob": 0}
    
    lambda_a = (team_a_adjusted["goals_scored"] + team_b_adjusted["goals_conceded"]) / max(team_b_weight, 0.1) / 2
    lambda_b = (team_b_adjusted["goals_scored"] + team_a_adjusted["goals_conceded"]) / max(team_a_weight, 0.1) / 2
    max_goals = 6
    prob_matrix = np.zeros((max_goals, max_goals))
    for i in range(max_goals):
        for j in range(max_goals):
            prob_matrix[i, j] = poisson.pmf(i, lambda_a) * poisson.pmf(j, lambda_b)
    
    most_likely_score = np.unravel_index(np.argmax(prob_matrix), prob_matrix.shape)
    win_prob = np.sum(prob_matrix[np.where(np.arange(max_goals)[:, None] > np.arange(max_goals)[None, :])])
    draw_prob = np.sum(np.diag(prob_matrix))
    loss_prob = np.sum(prob_matrix[np.where(np.arange(max_goals)[:, None] < np.arange(max_goals)[None, :])])
    ci_a = (poisson.ppf(0.075, lambda_a), poisson.ppf(0.925, lambda_a))
    ci_b = (poisson.ppf(0.075, lambda_b), poisson.ppf(0.925, lambda_b))
    total_goals = lambda_a + lambda_b
    over_2_5_prob = 1 - poisson.cdf(2, total_goals)

    # Converter prob_matrix para percentual
    prob_matrix_percent = prob_matrix * 100  # Multiplica por 100 para converter em percentual

    fig, ax = plt.subplots(figsize=(8, 6))
    sns.heatmap(prob_matrix_percent, annot=True, fmt=".1f", cmap="YlOrRd", 
                xticklabels=range(max_goals), yticklabels=range(max_goals),
                cbar_kws={'label': 'Probabilidade (%)'})
    ax.set_xlabel(f"Gols {team_b_name}")
    ax.set_ylabel(f"Gols {team_a_name}")
    ax.set_title("Distribuição de Probabilidade dos Placares (%)")
    plt.tight_layout()
    st.pyplot(fig)
    plt.close()

    return {
        "score": f"{most_likely_score[0]} x {most_likely_score[1]}",
        "probs": {"win": win_prob, "draw": draw_prob, "loss": loss_prob},
        "ci": {"team_a": ci_a, "team_b": ci_b},
        "total_goals_prob": over_2_5_prob
    }

# Função para buscar odds reais
def get_odds(fixture_id):
    if not API_KEY:
        st.error("Chave da API não configurada.")
        return []
    if not fixture_id:
        st.warning("Não foi possível encontrar um jogo futuro entre os times selecionados para buscar odds.")
        return []

    # Lista de mercados a buscar: Match Winner (1), Goals Over/Under (2), Both Teams To Score (8)
    market_ids = [1, 2, 8]  # IDs correspondentes aos mercados
    all_odds = []

    for bet_id in market_ids:
        url = f"{API_BASE_URL}/odds"
        params = {
            "fixture": fixture_id,
            "bet": bet_id
        }
        try:
            response = requests.get(url, headers=HEADERS, params=params)
            logger.debug(f"Resposta da API para odds (fixture {fixture_id}, bet {bet_id}): {response.status_code} - {response.text}")
            if response.status_code == 200:
                data = response.json()
                odds = data.get("response", [])
                if odds:
                    all_odds.extend(odds)
                else:
                    logger.debug(f"Nenhuma odd disponível para o mercado {bet_id} no fixture {fixture_id}")
            else:
                st.error(f"Erro ao buscar odds para mercado {bet_id}: {response.status_code} - {response.text}")
        except Exception as e:
            st.error(f"Erro ao buscar odds para mercado {bet_id}: {str(e)}")

    if not all_odds:
        st.warning("Nenhuma odd disponível para os mercados solicitados (Match Winner, Goals Over/Under, Both Teams To Score).")
    return all_odds

# Função para comparar odds e previsões
def compare_odds(predicted_stats, score_pred, odds):
    if not predicted_stats or score_pred["score"] == "N/A":
        st.warning("Não há previsões válidas para comparar com odds.")
        return []
    if not odds:
        # Já tratado em get_odds, mas mantemos por segurança
        st.warning("Nenhuma odd disponível para comparação.")
        return []

    comparison_data = []
    btts_prob = poisson.pmf(1, predicted_stats["goals_scored"]["team_a"]) * poisson.pmf(1, predicted_stats["goals_scored"]["team_b"])
    over_2_5_prob = score_pred["total_goals_prob"]
    under_2_5_prob = 1 - over_2_5_prob

    for market in odds:
        market_name = market.get("name", "Desconhecido")
        bets = market.get("bets", [])
        if not bets:
            continue
        for bet in bets:
            bet_name = bet.get("name", "Desconhecido")
            values = bet.get("values", [])
            if not values:
                continue
            if bet_name in ["Match Winner", "Both Teams To Score", "Goals Over/Under"]:
                for value in values:
                    odd = float(value.get("odd", 0))
                    if odd <= 0:
                        continue  # Ignora odds inválidas
                    implied_prob = 1 / odd
                    predicted_prob = 0
                    if bet_name == "Match Winner":
                        if value.get("value") == "Home":
                            predicted_prob = score_pred["probs"]["win"]
                        elif value.get("value") == "Draw":
                            predicted_prob = score_pred["probs"]["draw"]
                        elif value.get("value") == "Away":
                            predicted_prob = score_pred["probs"]["loss"]
                    elif bet_name == "Both Teams To Score":
                        if value.get("value") == "Yes":
                            predicted_prob = btts_prob
                        elif value.get("value") == "No":
                            predicted_prob = 1 - btts_prob
                    elif bet_name == "Goals Over/Under" and value.get("value") in ["Over 2.5", "Under 2.5"]:
                        if value.get("value") == "Over 2.5":
                            predicted_prob = over_2_5_prob
                        elif value.get("value") == "Under 2.5":
                            predicted_prob = under_2_5_prob

                    if predicted_prob > 0:  # Só adiciona se houver uma previsão válida
                        comparison_data.append({
                            "market": market_name,
                            "bet": bet_name,
                            "value": value.get("value", "Desconhecido"),
                            "odd": odd,
                            "implied_prob": implied_prob * 100,
                            "predicted_prob": predicted_prob * 100
                        })

    if not comparison_data:
        st.warning("Nenhuma odd válida para os mercados principais (Match Winner, Goals Over/Under, Both Teams To Score) foi encontrada ou os dados não atendem aos critérios de comparação.")
        return []

    return comparison_data

# Função para exportar resultados
def export_results(team_a, team_b, simple_a, adjusted_a, simple_b, adjusted_b, predicted_stats, confidence_intervals, score_pred, comparison_data, team_a_counts, team_b_counts, predicted_counts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Médias"
    ws.append(["Time", "Estatística", "Média Simples", "IC 85%", "Média Ajustada", "Nº Partidas"])
    
    z = norm.ppf(0.925)  # Para 85% de confiança
    for stat in simple_a.keys():
        # Aproximação de IC 85% para média simples usando Poisson para gols
        mean_a = simple_a[stat]
        mean_b = simple_b[stat]
        if stat == "goals_scored" or stat == "goals_conceded":
            std_a = np.sqrt(mean_a)
            std_b = np.sqrt(mean_b)
        else:
            std_a = mean_a / 2  # Aproximação arbitrária
            std_b = mean_b / 2
        error_a = std_a / np.sqrt(team_a_counts[stat]) if team_a_counts[stat] > 0 else 0
        error_b = std_b / np.sqrt(team_b_counts[stat]) if team_b_counts[stat] > 0 else 0
        ic_a = (max(0, mean_a - z * error_a), mean_a + z * error_a)
        ic_b = (max(0, mean_b - z * error_b), mean_b + z * error_b)
        
        ws.append([team_a["team"]["name"], stat, round(simple_a[stat], 1), f"[{round(ic_a[0], 1)}, {round(ic_a[1], 1)}]", round(adjusted_a[stat], 1), team_a_counts[stat]])
        ws.append([team_b["team"]["name"], stat, round(simple_b[stat], 1), f"[{round(ic_b[0], 1)}, {round(ic_b[1], 1)}]", round(adjusted_b[stat], 1), team_b_counts[stat]])

    ws = wb.create_sheet("Previsões")
    ws.append(["Estatística", f"{team_a['team']['name']} (Prev)", f"{team_a['team']['name']} (IC 85%)", f"{team_b['team']['name']} (Prev)", f"{team_b['team']['name']} (IC 85%)", "Nº Partidas"])
    for stat in predicted_stats.keys():
        ws.append([
            stat,
            round(predicted_stats[stat]["team_a"], 1),
            f"[{round(confidence_intervals[stat]['team_a'][0], 1)}, {round(confidence_intervals[stat]['team_a'][1], 1)}]",
            round(predicted_stats[stat]["team_b"], 1),
            f"[{round(confidence_intervals[stat]['team_b'][0], 1)}, {round(confidence_intervals[stat]['team_b'][1], 1)}]",
            predicted_counts[stat]
        ])
    ws = wb.create_sheet("Placar Provável")
    ws.append(["Placar", score_pred["score"]])
    ws.append(["Prob. Vitória", round(score_pred["probs"]["win"], 1)])
    ws.append(["Prob. Empate", round(score_pred["probs"]["draw"], 1)])
    ws.append(["Prob. Derrota", round(score_pred["probs"]["loss"], 1)])
    ws.append([f"IC Gols {team_a['team']['name']}", f"[{round(score_pred['ci']['team_a'][0], 1)}, {round(score_pred['ci']['team_a'][1], 1)}]"])
    ws.append([f"IC Gols {team_b['team']['name']}", f"[{round(score_pred['ci']['team_b'][0], 1)}, {round(score_pred['ci']['team_b'][1], 1)}]"])
    ws = wb.create_sheet("Odds e Probabilidades")
    ws.append(["Mercado", "Aposta", "Valor", "Odd", "Prob. Implícita (%)", "Prob. Prevista (%)"])
    for data in comparison_data:
        ws.append([
            data["market"],
            data["bet"],
            data["value"],
            round(data["odd"], 1),
            round(data["implied_prob"], 1),
            round(data["predicted_prob"], 1)
        ])
    filename = f"previsao_{team_a['team']['name']}_vs_{team_b['team']['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

# Função principal
def main():
    st.set_page_config(page_title="Previsão de Partidas de Futebol", layout="wide")
    st.title("Previsão Estatística de Partidas de Futebol")

    # Inicializar pesos dos adversários no session_state
    if "opponent_weights_a" not in st.session_state:
        st.session_state["opponent_weights_a"] = {}
    if "opponent_weights_b" not in st.session_state:
        st.session_state["opponent_weights_b"] = {}
    # Inicializar pontos Elo dos adversários no session_state
    if "opponent_elo_a" not in st.session_state:
        st.session_state["opponent_elo_a"] = {}
    if "opponent_elo_b" not in st.session_state:
        st.session_state["opponent_elo_b"] = {}

    # Upload do CSV com ratings Elo
    st.subheader("Carregar Ratings Elo dos Times")
    csv_file = st.file_uploader("Carregue o CSV com os ratings Elo (formato: Rank, Club / Country, Points, 1-yr change)", type="csv")
    if csv_file:
        try:
            ratings_df = pd.read_csv(csv_file)
            required_columns = ["Rank", "Club / Country", "Points", "1-yr change"]
            if all(col in ratings_df.columns for col in required_columns):
                ratings_df["Club / Country"] = ratings_df["Club / Country"].apply(reformat_club_country)
                st.session_state["ratings_df"] = ratings_df
                st.session_state["csv_name"] = csv_file.name
                st.success("CSV carregado com sucesso!")
                st.write("Dados carregados do CSV:")
                st.dataframe(ratings_df)
            else:
                st.error("O CSV não contém as colunas esperadas: Rank, Club / Country, Points, 1-yr change.")
        except Exception as e:
            st.error(f"Erro ao ler o CSV: {str(e)}")
    else:
        st.warning("Por favor, carregue um CSV para definir os pontos Elo dos times.")

    ratings_df = st.session_state.get("ratings_df", None)
    csv_name = st.session_state.get("csv_name", "Nenhum CSV carregado")

    tabs = st.tabs([
        "Home",
        "Seleção de Times",
        "Jogos Analisados",
        "Médias",
        "Estatísticas Previstas",
        "Placar Provável",
        "Odds x Previsão",
        "Exportar"
    ])

    with tabs[0]:
        st.header("Bem-vindo ao Prevelo")
        st.write("Este aplicativo utiliza os ratings Elo para definir os pesos dos times.")
        st.write(f"Fonte dos ratings: **{csv_name}**")

    with tabs[1]:
        st.subheader("Verificação da Chave da API")
        if st.button("Testar Chave"):
            if test_api_key():
                st.success("Chave da API está funcionando!")
            else:
                st.error("Chave da API inválida ou houve um erro.")

        col1, col2 = st.columns(2)
        with col1:
            team_a_name = st.text_input("Time A", placeholder="Digite o nome do time A")
            season_a = st.selectbox("Temporada Time A", list(range(2020, 2026)), index=5)
        with col2:
            team_b_name = st.text_input("Time B", placeholder="Digite o nome do time B")
            season_b = st.selectbox("Temporada Time B", list(range(2020, 2026)), index=5)

        campo_neutro = st.checkbox("Campo neutro (usar jogos gerais em vez de casa/fora)")
        st.session_state["campo_neutro"] = campo_neutro

        num_jogos = st.number_input("Número de jogos a analisar por time", min_value=1, max_value=20, value=10, step=1)
        st.session_state["num_jogos"] = num_jogos


        if st.button("Buscar Times"):
            if len(team_a_name) < 3 or len(team_b_name) < 3:
                st.error("O nome do time deve ter pelo menos 3 caracteres.")
            else:
                # Armazena season_a e season_b no session_state ao buscar os times
                st.session_state["season_a"] = season_a
                st.session_state["season_b"] = season_b
                teams_a = search_team(team_a_name)
                teams_b = search_team(team_b_name)
                st.session_state["teams_a"] = teams_a
                st.session_state["teams_b"] = teams_b
                if not teams_a:
                    st.error("Nenhum time encontrado para o Time A.")
                if not teams_b:
                    st.error("Nenhum time encontrado para o Time B.")
        if "teams_a" in st.session_state and "teams_b" in st.session_state and st.session_state["teams_a"] and st.session_state["teams_b"]:
            st.subheader("Selecione os Times")
            team_a_options = [f"{t['team']['name']} ({t['team']['country']})" for t in st.session_state["teams_a"]]
            team_b_options = [f"{t['team']['name']} ({t['team']['country']})" for t in st.session_state["teams_b"]]
            team_a_selected = st.selectbox("Time A", team_a_options, key="team_a_select")
            team_b_selected = st.selectbox("Time B", team_b_options, key="team_b_select")

            # Selecionar os times
            st.session_state["team_a"] = next(t for t in st.session_state["teams_a"] if f"{t['team']['name']} ({t['team']['country']})" == team_a_selected)
            st.session_state["team_b"] = next(t for t in st.session_state["teams_b"] if f"{t['team']['name']} ({t['team']['country']})" == team_b_selected)

            # Obter pontos Elo e peso iniciais
            team_a_elo, team_a_weight = get_team_elo_and_weight(st.session_state["team_a"]["team"]["name"], ratings_df)
            team_b_elo, team_b_weight = get_team_elo_and_weight(st.session_state["team_b"]["team"]["name"], ratings_df)

            # Mostrar pontos Elo iniciais e permitir ajuste
            col_a, col_b = st.columns(2)
            with col_a:
                initial_elo_a = team_a_elo if team_a_elo is not None else 1600
                if team_a_elo is None:
                    st.warning(f"Pontos Elo não identificados para {team_a_selected}.")
                st.write(f"Pontos Elo iniciais do {team_a_selected}: {initial_elo_a}")
                team_a_elo_adjusted = st.number_input(
                    f"Ajustar pontos Elo do {team_a_selected}",
                    min_value=0,
                    value=int(initial_elo_a),
                    step=1,
                    key="team_a_elo_adjusted"
                )
                team_a_weight_adjusted = calculate_weight_from_elo(team_a_elo_adjusted)
                st.write(f"Peso calculado: {team_a_weight_adjusted:.2f}")

            with col_b:
                initial_elo_b = team_b_elo if team_b_elo is not None else 1600
                if team_b_elo is None:
                    st.warning(f"Pontos Elo não identificados para {team_b_selected}.")
                st.write(f"Pontos Elo iniciais do {team_b_selected}: {initial_elo_b}")
                team_b_elo_adjusted = st.number_input(
                    f"Ajustar pontos Elo do {team_b_selected}",
                    min_value=0,
                    value=int(initial_elo_b),
                    step=1,
                    key="team_b_elo_adjusted"
                )
                team_b_weight_adjusted = calculate_weight_from_elo(team_b_elo_adjusted)
                st.write(f"Peso calculado: {team_b_weight_adjusted:.2f}")

            if st.button("Confirmar Seleção"):
                # Armazenar os pontos Elo ajustados e os pesos calculados
                st.session_state["team_a_elo"] = team_a_elo_adjusted
                st.session_state["team_b_elo"] = team_b_elo_adjusted
                st.session_state["team_a_weight"] = team_a_weight_adjusted
                st.session_state["team_b_weight"] = team_b_weight_adjusted
                st.success(f"Times selecionados com sucesso! Pesos ajustados: {team_a_selected}: {st.session_state['team_a_weight']:.2f}, {team_b_selected}: {st.session_state['team_b_weight']:.2f}")

    with tabs[2]:
        if "team_a" in st.session_state and "team_b" in st.session_state:
            team_a_id = st.session_state["team_a"]["team"]["id"]
            team_b_id = st.session_state["team_b"]["team"]["id"]
            season_a = st.session_state.get("season_a", 2025)
            season_b = st.session_state.get("season_b", 2025)
            games_a = get_team_games(
                team_a_id,
                season_a,
                home=True,
                limit=st.session_state.get("num_jogos", 10),
                neutral=st.session_state.get("campo_neutro", False)
            )

            games_b = get_team_games(
                team_b_id,
                season_b,
                home=False,
                limit=st.session_state.get("num_jogos", 10),
                neutral=st.session_state.get("campo_neutro", False)
            )


            ratings_df = st.session_state.get("ratings_df", None)

            # Botão no início da aba para recalcular
            if (games_a or games_b) and st.button("Confirmar Pesos dos Adversários e Calcular Médias e Estatísticas"):
                # Limpar estados anteriores relacionados a médias e previsões
                for key in [
                    "simple_a", "adjusted_a", "team_a_counts", "team_a_raw_stats", "team_a_raw_adjusted",
                    "simple_b", "adjusted_b", "team_b_counts", "team_b_raw_stats", "team_b_raw_adjusted",
                    "predicted_stats", "confidence_intervals", "predicted_counts", "score_pred", "comparison_data"
                ]:
                    if key in st.session_state:
                        del st.session_state[key]

                st.success("Médias e estatísticas foram recalculadas com os novos pesos. Verifique as abas 'Médias', 'Estatísticas Previstas' e 'Placar Provável'.")

            if games_a:
                st.write(f"Jogos do Time A ({st.session_state['team_a']['team']['name']} - Mandante):")
                for game in games_a:
                    game_date = datetime.strptime(game["fixture"]["date"], "%Y-%m-%dT%H:%M:%S+00:00")
                    formatted_date = game_date.strftime("%d/%m/%Y %H:%M")
                    home_team = game["teams"]["home"]["name"]
                    away_team = game["teams"]["away"]["name"]
                    home_goals = game["goals"]["home"] if game["goals"]["home"] is not None else 0
                    away_goals = game["goals"]["away"] if game["goals"]["away"] is not None else 0
                    league_name = game["league"]["name"]
                    stats = get_game_stats(game["fixture"]["id"])
                    has_stats = bool(stats and any(stat["team"]["id"] == team_a_id for stat in stats))
                    
                    # Corrigir se Time A for mandante ou visitante neste jogo
                    if game["teams"]["home"]["id"] == team_a_id:
                        opponent_name = game["teams"]["away"]["name"]
                    else:
                        opponent_name = game["teams"]["home"]["name"]
                    opponent_elo, opponent_weight = get_team_elo_and_weight(opponent_name, ratings_df)

                    
                    # Permitir ajuste dos pontos Elo do adversário
                    fixture_id = str(game["fixture"]["id"])
                    if fixture_id not in st.session_state["opponent_elo_a"]:
                        st.session_state["opponent_elo_a"][fixture_id] = opponent_elo if opponent_elo is not None else 1600

                    # Aviso se pontos Elo não identificados
                    if opponent_elo is None:
                        st.warning(f"Pontos Elo não identificados para o adversário {opponent_name}.")

                    opponent_elo_adjusted = st.number_input(
                        f"Pontos Elo do Adversário ({opponent_name})",
                        min_value=0,
                        value=int(st.session_state["opponent_elo_a"][fixture_id]),
                        step=1,
                        key=f"opponent_elo_a_{fixture_id}"
                    )
                    st.session_state["opponent_elo_a"][fixture_id] = opponent_elo_adjusted
                    opponent_weight_adjusted = calculate_weight_from_elo(opponent_elo_adjusted)
                    st.session_state["opponent_weights_a"][fixture_id] = opponent_weight_adjusted

                    title_suffix = " (SEM ESTATÍSTICAS)" if not has_stats else ""
                    title = f"{home_team} {home_goals} x {away_goals} {away_team} - {formatted_date} ({league_name}) - Peso Adversário: {opponent_weight_adjusted:.2f}{title_suffix}"
                    with st.expander(title):
                        if has_stats:
                            data = []
                            for stat in stats:
                                if stat["team"]["id"] == team_a_id:
                                    for s in stat["statistics"]:
                                        value = s["value"]
                                        if s["type"].lower() == "ball possession" and isinstance(value, str):
                                            value = value.replace("%", "").strip()
                                            value = float(value) if value else 0.0
                                        else:
                                            try:
                                                value = float(value) if value is not None else 0.0
                                            except (ValueError, TypeError):
                                                value = 0.0
                                        data.append([s["type"], value])
                            df_stats = pd.DataFrame(data, columns=["Estatística", "Valor"])
                            df_stats["Valor"] = df_stats["Valor"].round(1)
                            st.dataframe(df_stats)
                        else:
                            st.write("Nenhuma estatística disponível para este jogo.")
            else:
                st.warning(f"Nenhum jogo finalizado encontrado para {st.session_state['team_a']['team']['name']} na temporada {season_a}. Tente outra temporada, como 2024.")
            
            if games_b:
                st.write(f"Jogos do Time B ({st.session_state['team_b']['team']['name']} - Visitante):")
                for game in games_b:
                    game_date = datetime.strptime(game["fixture"]["date"], "%Y-%m-%dT%H:%M:%S+00:00")
                    formatted_date = game_date.strftime("%d/%m/%Y %H:%M")
                    home_team = game["teams"]["home"]["name"]
                    away_team = game["teams"]["away"]["name"]
                    home_goals = game["goals"]["home"] if game["goals"]["home"] is not None else 0
                    away_goals = game["goals"]["away"] if game["goals"]["away"] is not None else 0
                    league_name = game["league"]["name"]
                    stats = get_game_stats(game["fixture"]["id"])
                    has_stats = bool(stats and any(stat["team"]["id"] == team_b_id for stat in stats))
                    
                    # Corrigir se Time B for mandante ou visitante neste jogo
                    if game["teams"]["home"]["id"] == team_b_id:
                        opponent_name = game["teams"]["away"]["name"]
                    else:
                        opponent_name = game["teams"]["home"]["name"]
                    opponent_elo, opponent_weight = get_team_elo_and_weight(opponent_name, ratings_df)

                    
                    # Permitir ajuste dos pontos Elo do adversário
                    fixture_id = str(game["fixture"]["id"])
                    if fixture_id not in st.session_state["opponent_elo_b"]:
                        st.session_state["opponent_elo_b"][fixture_id] = opponent_elo if opponent_elo is not None else 1600

                    # Aviso se pontos Elo não identificados
                    if opponent_elo is None:
                        st.warning(f"Pontos Elo não identificados para o adversário {opponent_name}.")

                    opponent_elo_adjusted = st.number_input(
                        f"Pontos Elo do Adversário ({opponent_name})",
                        min_value=0,
                        value=int(st.session_state["opponent_elo_b"][fixture_id]),
                        step=1,
                        key=f"opponent_elo_b_{fixture_id}"
                    )
                    st.session_state["opponent_elo_b"][fixture_id] = opponent_elo_adjusted
                    opponent_weight_adjusted = calculate_weight_from_elo(opponent_elo_adjusted)
                    st.session_state["opponent_weights_b"][fixture_id] = opponent_weight_adjusted

                    title_suffix = " (SEM ESTATÍSTICAS)" if not has_stats else ""
                    title = f"{home_team} {home_goals} x {away_goals} {away_team} - {formatted_date} ({league_name}) - Peso Adversário: {opponent_weight_adjusted:.2f}{title_suffix}"
                    with st.expander(title):
                        if has_stats:
                            data = []
                            for stat in stats:
                                if stat["team"]["id"] == team_b_id:
                                    for s in stat["statistics"]:
                                        value = s["value"]
                                        if s["type"].lower() == "ball possession" and isinstance(value, str):
                                            value = value.replace("%", "").strip()
                                            value = float(value) if value else 0.0
                                        else:
                                            try:
                                                value = float(value) if value is not None else 0.0
                                            except (ValueError, TypeError):
                                                value = 0.0
                                        data.append([s["type"], value])
                            df_stats = pd.DataFrame(data, columns=["Estatística", "Valor"])
                            df_stats["Valor"] = df_stats["Valor"].round(1)
                            st.dataframe(df_stats)
                        else:
                            st.write("Nenhuma estatística disponível para este jogo.")
            else:
                st.warning(f"Nenhum jogo finalizado encontrado para {st.session_state['team_b']['team']['name']} na temporada {season_b}. Tente outra temporada, como 2024.")

        else:
            st.info("Selecione os times na aba 'Seleção de Times' para ver os jogos.")

    with tabs[3]:
        if "team_a" in st.session_state and "team_b" in st.session_state:
            team_a_id = st.session_state["team_a"]["team"]["id"]
            team_b_id = st.session_state["team_b"]["team"]["id"]
            season_a = st.session_state.get("season_a", 2025)
            season_b = st.session_state.get("season_b", 2025)
            team_a_weight = st.session_state.get("team_a_weight", 0.8)
            team_b_weight = st.session_state.get("team_b_weight", 0.8)
            games_a = get_team_games(
                team_a_id,
                season_a,
                home=True,
                neutral=st.session_state.get("campo_neutro", False)
            )

            games_b = get_team_games(
                team_b_id,
                season_b,
                home=False,
                neutral=st.session_state.get("campo_neutro", False)
            )

            if games_a and games_b:
                simple_a, adjusted_a, team_a_counts, _, team_a_raw_stats, team_a_raw_adjusted = calculate_averages(
                    games_a, team_a_id, season_a, team_a_weight, st.session_state["opponent_weights_a"]
                )
                simple_b, adjusted_b, team_b_counts, _, team_b_raw_stats, team_b_raw_adjusted = calculate_averages(
                    games_b, team_b_id, season_b, team_b_weight, st.session_state["opponent_weights_b"]
                )
                st.session_state["simple_a"] = simple_a
                st.session_state["adjusted_a"] = adjusted_a
                st.session_state["simple_b"] = simple_b
                st.session_state["adjusted_b"] = adjusted_b
                st.session_state["team_a_counts"] = team_a_counts
                st.session_state["team_b_counts"] = team_b_counts
                st.session_state["team_a_raw_stats"] = team_a_raw_stats
                st.session_state["team_a_raw_adjusted"] = team_a_raw_adjusted
                st.session_state["team_b_raw_stats"] = team_b_raw_stats
                st.session_state["team_b_raw_adjusted"] = team_b_raw_adjusted

                # Calcular IC 85% para médias simples
                z = norm.ppf(0.925)  # Para 85% de confiança
                ic_a = {}
                ic_b = {}
                for stat in simple_a.keys():
                    mean_a = simple_a[stat]
                    mean_b = simple_b[stat]
                    if stat == "goals_scored" or stat == "goals_conceded":
                        std_a = np.sqrt(mean_a)
                        std_b = np.sqrt(mean_b)
                    else:
                        std_a = mean_a / 2  # Aproximação
                        std_b = mean_b / 2
                    error_a = std_a / np.sqrt(team_a_counts[stat]) if team_a_counts[stat] > 0 else 0
                    error_b = std_b / np.sqrt(team_b_counts[stat]) if team_b_counts[stat] > 0 else 0
                    ic_a[stat] = (max(0, mean_a - z * error_a), mean_a + z * error_a)
                    ic_b[stat] = (max(0, mean_b - z * error_b), mean_b + z * error_b)

                df_a = pd.DataFrame({
                    "Estatística": simple_a.keys(),
                    "Média Simples": [round(v, 1) for v in simple_a.values()],
                    "IC 85%": [f"[{round(ic_a[k][0], 1)}, {round(ic_a[k][1], 1)}]" for k in simple_a.keys()],
                    "Média Ajustada": [round(v, 1) for v in adjusted_a.values()],
                    "Nº Partidas": [team_a_counts[k] for k in simple_a.keys()]
                })
                df_b = pd.DataFrame({
                    "Estatística": simple_b.keys(),
                    "Média Simples": [round(v, 1) for v in simple_b.values()],
                    "IC 85%": [f"[{round(ic_b[k][0], 1)}, {round(ic_b[k][1], 1)}]" for k in simple_b.keys()],
                    "Média Ajustada": [round(v, 1) for v in adjusted_b.values()],
                    "Nº Partidas": [team_b_counts[k] for k in simple_b.keys()]
                })
                st.write(f"Médias do Time A ({st.session_state['team_a']['team']['name']}):")
                st.dataframe(df_a)
                st.write(f"Peso do Time A: {team_a_weight:.2f}")
                st.write(f"Médias do Time B ({st.session_state['team_b']['team']['name']}):")
                st.dataframe(df_b)
                st.write(f"Peso do Time B: {team_b_weight:.2f}")
            else:
                st.warning("Não foi possível calcular médias. Verifique se há jogos finalizados na aba 'Jogos Analisados'.")
        else:
            st.info("Selecione os times na aba 'Seleção de Times' para calcular as médias.")

    with tabs[4]:
        if "simple_a" in st.session_state:
            team_a_weight = st.session_state.get("team_a_weight", 0.8)
            team_b_weight = st.session_state.get("team_b_weight", 0.8)
            predicted_stats, confidence_intervals, predicted_counts = predict_stats(
                st.session_state["simple_a"], st.session_state["adjusted_a"],
                st.session_state["simple_b"], st.session_state["adjusted_b"],
                st.session_state["team_a_counts"], st.session_state["team_b_counts"],
                team_a_weight, team_b_weight
            )
            st.session_state["predicted_stats"] = predicted_stats
            st.session_state["confidence_intervals"] = confidence_intervals
            st.session_state["predicted_counts"] = predicted_counts
            if predicted_stats:
                df_pred = pd.DataFrame([
                    {
                        "Estatística": stat,
                        f"{st.session_state['team_a']['team']['name']}": round(pred["team_a"], 1),
                        f"{st.session_state['team_a']['team']['name']} (IC 85%)": f"[{round(confidence_intervals[stat]['team_a'][0], 1)}, {round(confidence_intervals[stat]['team_a'][1], 1)}]",
                        f"{st.session_state['team_b']['team']['name']}": round(pred["team_b"], 1),
                        f"{st.session_state['team_b']['team']['name']} (IC 85%)": f"[{round(confidence_intervals[stat]['team_b'][0], 1)}, {round(confidence_intervals[stat]['team_b'][1], 1)}]",
                        "Nº Partidas": predicted_counts[stat]
                    }
                    for stat, pred in predicted_stats.items()
                ])
                st.dataframe(df_pred)
            else:
                st.warning("Não foi possível gerar previsões devido à falta de dados de gols.")
        else:
            st.info("Calcule as médias na aba 'Médias' para gerar previsões.")

    with tabs[5]:
        if "adjusted_a" in st.session_state:
            team_a_weight = st.session_state.get("team_a_weight", 0.8)
            team_b_weight = st.session_state.get("team_b_weight", 0.8)
            score_pred = predict_score(
                st.session_state["adjusted_a"], 
                st.session_state["adjusted_b"],
                team_a_weight, 
                team_b_weight,
                st.session_state["team_a"]["team"]["name"],
                st.session_state["team_b"]["team"]["name"]
            )
            st.session_state["score_pred"] = score_pred
            if score_pred["score"] != "N/A":
                st.write(f"Placar mais provável: {score_pred['score']}")
                st.write(f"Probabilidade de Vitória: {score_pred['probs']['win']*100:.1f}%")
                st.write(f"Probabilidade de Empate: {score_pred['probs']['draw']*100:.1f}%")
                st.write(f"Probabilidade de Derrota: {score_pred['probs']['loss']*100:.1f}%")
                st.write(f"Intervalo de Confiança (Gols {st.session_state['team_a']['team']['name']}): [{score_pred['ci']['team_a'][0]:.1f}, {score_pred['ci']['team_a'][1]:.1f}]")
                st.write(f"Intervalo de Confiança (Gols {st.session_state['team_b']['team']['name']}): [{score_pred['ci']['team_b'][0]:.1f}, {score_pred['ci']['team_b'][1]:.1f}]")
            else:
                st.warning("Não foi possível prever o placar devido à falta de dados de gols.")
        else:
            st.info("Calcule as médias na aba 'Médias' para prever o placar.")

    with tabs[6]:
        if "team_a" in st.session_state and "team_b" in st.session_state and "season_a" in st.session_state and "predicted_stats" in st.session_state:
            team_a_id = st.session_state["team_a"]["team"]["id"]
            team_b_id = st.session_state["team_b"]["team"]["id"]
            season = st.session_state["season_a"]
            fixture_id = find_next_fixture(team_a_id, team_b_id, season)
            odds = get_odds(fixture_id)
            comparison_data = compare_odds(st.session_state["predicted_stats"], st.session_state["score_pred"], odds)
            st.session_state["comparison_data"] = comparison_data
            if comparison_data:
                df_odds = pd.DataFrame(comparison_data)
                df_odds["implied_prob"] = df_odds["implied_prob"].round(1)
                df_odds["predicted_prob"] = df_odds["predicted_prob"].round(1)
                df_odds["odd"] = df_odds["odd"].round(1)
                st.dataframe(df_odds)
            else:
                st.warning("Nenhuma odd válida disponível para comparação nos mercados principais.")
        else:
            st.info("Selecione os times na aba 'Seleção de Times' e complete as previsões nas abas anteriores para comparar odds.")

    with tabs[7]:
        if "team_a" in st.session_state and "predicted_stats" in st.session_state and st.session_state["predicted_stats"]:
            if st.button("Exportar Resultados"):
                filename = export_results(
                    st.session_state["team_a"], st.session_state["team_b"],
                    st.session_state["simple_a"], st.session_state["adjusted_a"],
                    st.session_state["simple_b"], st.session_state["adjusted_b"],
                    st.session_state["predicted_stats"], st.session_state["confidence_intervals"],
                    st.session_state["score_pred"], st.session_state["comparison_data"],
                    st.session_state["team_a_counts"], st.session_state["team_b_counts"],
                    st.session_state["predicted_counts"]
                )
                with open(filename, "rb") as f:
                    st.download_button("Baixar .xlsx", f, file_name=filename)
                st.success("Resultados exportados com sucesso!")
        else:
            st.info("Complete as previsões nas abas anteriores para exportar os resultados.")

if __name__ == "__main__":
    main()